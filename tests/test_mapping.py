"""Testes: regra matemática, parsing de cartas, suporte a múltiplos idiomas e
checagem cruzada com a planilha original tabela_baralho_bip39.xlsx (quando
presente em docs/)."""

import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent / "src"))

from bip39_baralho import deck, mapping  # noqa: E402
from bip39_baralho.idiomas import IDIOMAS, carregar_wordlist  # noqa: E402


def test_wordlist_padrao_tem_2048_palavras_unicas():
    assert len(mapping.WORDLIST) == 2048
    assert len(set(mapping.WORDLIST)) == 2048


def test_todos_os_idiomas_tem_2048_palavras_unicas():
    for idioma in IDIOMAS:
        palavras = carregar_wordlist(idioma)
        assert len(palavras) == 2048, idioma
        assert len(set(palavras)) == 2048, idioma


def test_parse_card_variacoes():
    assert deck.parse_card("A♣") == 0
    assert deck.parse_card("AC") == 0
    assert deck.parse_card("2C") == 1
    assert deck.parse_card("KS") == 51
    assert deck.parse_card("10-Copas") == 26 + 9
    assert deck.parse_card("Q Espadas") == 39 + 11


def test_primeira_combinacao_e_abandon_em_ingles():
    r = mapping.combinar(0, 1)  # A♣ -> 2♣
    assert r.n == 0
    assert r.valido is True
    assert r.palavra == "abandon"


def test_primeira_combinacao_em_portugues():
    r = mapping.combinar(0, 1, idioma="portuguese")
    assert r.n == 0
    assert r.valido is True
    assert r.palavra == "abacate"


def test_primeira_combinacao_em_espanhol():
    r = mapping.combinar(0, 1, idioma="spanish")
    assert r.valido is True
    assert r.palavra == "ábaco"


def test_idioma_desconhecido_gera_erro():
    try:
        mapping.combinar(0, 1, idioma="klingon")
        assert False, "deveria ter levantado ValueError"
    except ValueError:
        pass


def test_ultima_combinacao_valida():
    mapa, _ = mapping.gerar_tabela_completa()
    assert len(mapa) == 2048
    assert mapa[-1].n == 2047
    assert mapa[-1].palavra == mapping.WORDLIST[2047]


def test_descartes_tem_604_combinacoes():
    _, descartes = mapping.gerar_tabela_completa()
    assert len(descartes) == 604
    assert all(r.n >= mapping.DISCARD_MIN for r in descartes)


def test_combinacao_invalida_com_cartas_iguais():
    try:
        mapping.combinar(5, 5)
        assert False, "deveria ter levantado ValueError"
    except ValueError:
        pass


def test_tabela_completa_bate_com_planilha_original_se_disponivel():
    xlsx_ref = pathlib.Path(__file__).resolve().parent.parent / "docs" / "tabela_baralho_bip39.xlsx"
    if not xlsx_ref.exists():
        return  # planilha de referência não incluída neste checkout; pula o teste
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_ref, data_only=True)
    xl_mapa = list(wb["Mapa BIP39"].iter_rows(min_row=2, values_only=True))
    xl_desc = list(wb["Descartes"].iter_rows(min_row=2, values_only=True))

    mapa, descartes = mapping.gerar_tabela_completa()
    assert len(mapa) == len(xl_mapa)
    assert len(descartes) == len(xl_desc)

    for r, x in zip(mapa, xl_mapa):
        gerado = (r.n, r.palavra, format(r.n, "011b"), r.a, deck.card_code(r.a), deck.card_name(r.a),
                  r.b_ajustado, r.b, deck.card_code(r.b), deck.card_name(r.b), r.padrao, r.n)
        assert gerado == tuple(x)

    for r, x in zip(descartes, xl_desc):
        gerado = (r.n, r.a, deck.card_code(r.a), deck.card_name(r.a), r.b_ajustado,
                  r.b, deck.card_code(r.b), deck.card_name(r.b), r.padrao, "DESCARTAR")
        assert gerado == tuple(x)
